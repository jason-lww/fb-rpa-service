from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook


PHONE_HEADERS = ("手机号", "手机号码", "号码", "phone", "Phone", "number", "Number")
STATUS_HEADERS = ("状态", "status", "Status")
ACCOUNT_HEADERS = ("账套", "账号账套", "投放状态", "account", "Account")
BLOCKED_BEFORE_BIND = "绑定前封号"
READY_FOR_DELIVERY = "待投放"
DELIVERY_DONE = "投放完成"
REMOVED = "已移除"


@dataclass
class PhonePoolRow:
    sheet_name: str
    row_index: int
    phone: str
    status: str
    phone_col: int
    status_col: Optional[int]
    account_col: Optional[int]


def pick_next_bindable_phone(file_path: str, expected_phone: Optional[str] = None) -> PhonePoolRow:
    for row in _iter_phone_rows(file_path):
        if expected_phone and _normalize_phone(row.phone) != _normalize_phone(expected_phone):
            continue
        if row.status.strip() and row.status.strip() != BLOCKED_BEFORE_BIND:
            continue
        if row.status.strip() == BLOCKED_BEFORE_BIND:
            continue
        return row
    raise ValueError("号码池中没有可绑定号码")


def list_delivery_completed(file_path: str) -> list[PhonePoolRow]:
    return [row for row in _iter_phone_rows(file_path) if row.status.strip() == DELIVERY_DONE]


def mark_ready_for_delivery(file_path: str, row: PhonePoolRow):
    _write_row_value(file_path, row, row.account_col or row.status_col, READY_FOR_DELIVERY)


def mark_removed(file_path: str, row: PhonePoolRow):
    _write_row_value(file_path, row, row.status_col or row.account_col, REMOVED)


def _iter_phone_rows(file_path: str) -> Iterable[PhonePoolRow]:
    workbook = load_workbook(Path(file_path))
    for worksheet in reversed(workbook.worksheets):
        header_map = _read_headers(worksheet)
        phone_col = _find_col(header_map, PHONE_HEADERS)
        if not phone_col:
            continue
        status_col = _find_col(header_map, STATUS_HEADERS)
        account_col = _find_col(header_map, ACCOUNT_HEADERS)
        for row_index in range(2, worksheet.max_row + 1):
            phone = _cell_text(worksheet.cell(row=row_index, column=phone_col).value)
            if not phone:
                continue
            status = _cell_text(worksheet.cell(row=row_index, column=status_col).value) if status_col else ""
            yield PhonePoolRow(
                sheet_name=worksheet.title,
                row_index=row_index,
                phone=phone,
                status=status,
                phone_col=phone_col,
                status_col=status_col,
                account_col=account_col,
            )


def _write_row_value(file_path: str, row: PhonePoolRow, column: Optional[int], value: str):
    if not column:
        raise ValueError("号码池缺少可更新的状态/账套列")
    workbook = load_workbook(Path(file_path))
    worksheet = workbook[row.sheet_name]
    worksheet.cell(row=row.row_index, column=column).value = value
    workbook.save(Path(file_path))


def _read_headers(worksheet) -> dict[str, int]:
    headers = {}
    for cell in worksheet[1]:
        text = _cell_text(cell.value)
        if text:
            headers[text] = cell.column
    return headers


def _find_col(header_map: dict[str, int], aliases: tuple[str, ...]) -> Optional[int]:
    for alias in aliases:
        if alias in header_map:
            return header_map[alias]
    return None


def _cell_text(value) -> str:
    return "" if value is None else str(value).strip()


def _normalize_phone(phone: str) -> str:
    return "".join(ch for ch in phone if ch.isdigit())
